# -*- coding: utf-8 -*-
##############################################################################
#
# OpenERP, Open Source Management Solution, third party addon
# Copyright (C) 2004-2016  Vertel AB (<http://vertel.se>).
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

from openerp import api, models, fields, _

from openpyxl import Workbook
import cStringIO
import base64

import logging
_logger = logging.getLogger(__name__)

class accounting_report(models.TransientModel):
    _inherit = "accounting.report"
    _description = "Accounting Report"

    state = fields.Selection(selection=[('get','get'),('put','put')],default=None)
    excel_file = fields.Binary()
    filename = fields.Char(default="financial_report.xls")

    #~ def check_excel(self, cr, uid, ids, context=None):
        #~ if context is None:
            #~ context = {}
        #~ res = super(accounting_report, self).check_report(cr, uid, ids, context=context)
        #~ data = {}
        #~ data['form'] = self.read(cr, uid, ids, ['account_report_id', 'date_from_cmp',  'date_to_cmp',  'fiscalyear_id_cmp', 'journal_ids', 'period_from_cmp', 'period_to_cmp',  'filter_cmp',  'chart_account_id', 'target_move'], context=context)[0]
        #~ for field in ['fiscalyear_id_cmp', 'chart_account_id', 'period_from_cmp', 'period_to_cmp', 'account_report_id']:
            #~ if isinstance(data['form'][field], tuple):
                #~ data['form'][field] = data['form'][field][0]
        #~ comparison_context = self._build_comparison_context(cr, uid, ids, data, context=context)
        #~ res['data']['form']['comparison_context'] = comparison_context
        #~ return res

    @api.multi
    def check_excel(self):
        for s in self:
            wb = Workbook()
            ws = wb.active
            #raise Warning(data['form']['account_report_id'][0])
            ws.title = s.account_report_id.name
            ws.merge_cells(start_row=1,end_row=1,start_column=1,end_column=6)
            ws.cell(row = 1, column = 1).value = ws.title
            r = 2
            for line in s.get_lines():
                if line['type'] == 'report' and line['account_type'] == 'view':
                    c = 6
                    n = 1
                elif line['type'] == 'report':
                    c = 5
                    n = 2
                else:
                    c = 4
                    n = 3
                ws.merge_cells(start_row=r,end_row=r,start_column=n,end_column=3)
                ws.cell(row = r, column = n).value = line['name']
                ws.cell(row = r, column = c).value = line['balance']
                #~ ws.cell(row = r, column = 7).value = line['account_type']
                #~ ws.cell(row = r, column = 8).value = line['type']
                #~ ws.cell(row = r, column = 9).value = line['level']
                #{'account_type': False, 'balance': 0.0, 'type': 'report', 'name': u'2.2 F\xf6rskott avseende immateriella anl\xe4ggningstillg\xe5ngar', 'level': 4}, {'account_type': 'view', 'balance': 0.0, 'type': 'report', 'name': u'Materiella anl\xe4ggningstillg\xe5ngar', 'level': 3}
                r += 1
            out = cStringIO.StringIO()
            wb.save(out)
            s.excel_file=base64.b64encode(out.getvalue())
            s.state = 'get'
            out.close()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'accounting.report',
            'view_mode': 'form',
            'view_type': 'form',
            'view_id': self.env.ref('account_financial_excel.accounting_report_excel_result').id,
            'res_id': self[0].id,
            'views': [(False, 'form')],
            'target': 'new',
        }


        return res


    def _print_report(self, cr, uid, ids, data, context=None):
        data['form'].update(self.read(cr, uid, ids, ['date_from_cmp',  'debit_credit', 'date_to_cmp',  'fiscalyear_id_cmp', 'period_from_cmp', 'period_to_cmp',  'filter_cmp', 'account_report_id', 'enable_filter', 'label_filter','target_move'], context=context)[0])
        data['form']
        wb = Workbook()
        ws = wb.active
        #raise Warning(data['form']['account_report_id'][0])
        ws.title = self.pool.get('account.financial.report').browse(cr, uid,data['form']['account_report_id'][0]).name 
        ws.merge_cells(start_row=1,end_row=1,start_column=1,end_column=6)
        ws.cell(row = 1, column = 1).value = ws.title
        r = 2
        for line in self.get_lines(cr,uid,data):
            if line['type'] == 'report' and line['account_type'] == 'view':
                c = 6
                n = 1
            elif line['type'] == 'report':
                c = 5
                n = 2
            else:
                c = 4
                n = 3
            ws.merge_cells(start_row=r,end_row=r,start_column=n,end_column=3)
            ws.cell(row = r, column = n).value = line['name']
            ws.cell(row = r, column = c).value = line['balance']
            #~ ws.cell(row = r, column = 7).value = line['account_type']
            #~ ws.cell(row = r, column = 8).value = line['type']
            #~ ws.cell(row = r, column = 9).value = line['level']
            #{'account_type': False, 'balance': 0.0, 'type': 'report', 'name': u'2.2 F\xf6rskott avseende immateriella anl\xe4ggningstillg\xe5ngar', 'level': 4}, {'account_type': 'view', 'balance': 0.0, 'type': 'report', 'name': u'Materiella anl\xe4ggningstillg\xe5ngar', 'level': 3}
            r += 1
        out = cStringIO.StringIO()
        wb.save(out)
        self.write(cr,uid,ids[0],{'excel_file': base64.b64encode(out.getvalue()), 'state': 'get'})
        out.close()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'accounting.report',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': ids[0],
            'views': [(False, 'form')],
            'target': 'new',
        }

    
    def get_lines(self):        
        lines = []
        report = self.account_report_id.id
        ids2 = self.pool.get('account.financial.report')._get_children_by_order(self._cr,self._uid,[report])
        for report in self.env['account.financial.report'].browse(ids2):
            vals = {
                'name': report.name,
                'balance': report.balance * report.sign or 0.0,
                'type': 'report',
                'level': bool(report.style_overwrite) and report.style_overwrite or report.level,
                'account_type': report.type =='sum' and 'view' or False, #used to underline the financial report balances
            }
            if self.debit_credit:
                vals['debit'] = report.debit
                vals['credit'] = report.credit
            lines.append(vals)
            account_ids = []
            if report.display_detail == 'no_detail':
                #the rest of the loop is used to display the details of the financial report, so it's not needed here.
                continue
            if report.type == 'accounts' and report.account_ids:
                account_ids = self.pool.get('account.account')._get_children_and_consol(self._cr,self._uid,[x.id for x in report.account_ids])
            elif report.type == 'account_type' and report.account_type_ids:
                account_ids = [a.id for a in self.env['account.account'].search([('user_type','in', [x.id for x in report.account_type_ids])])]
            if account_ids:
                for account in self.env['account.account'].browse(account_ids):
                    #if there are accounts to display, we add them to the lines with a level equals to their level in
                    #the COA + 1 (to avoid having them with a too low level that would conflicts with the level of data
                    #financial reports for Assets, liabilities...)
                    if report.display_detail == 'detail_flat' and account.type == 'view':
                        continue
                    flag = False
                    vals = {
                        'name': '%s %s' % (account.code,account.name),
                        'balance':  account.balance != 0 and account.balance * report.sign or account.balance,
                        'type': 'account',
                        'level': report.display_detail == 'detail_with_hierarchy' and min(account.level + 1,6) or 6, #account.level + 1
                        'account_type': account.type,
                    }

                    if self.debit_credit:
                        vals['debit'] = account.debit
                        vals['credit'] = account.credit
                    if not self.pool.get('res.currency').is_zero(self._cr, self._uid, account.company_id.currency_id, vals['balance']):
                        flag = True
                    if flag:
                        lines.append(vals)
        return lines
    
    def _build_contexts(self):
        result = {}
        result['fiscalyear'] = self.fiscalyear_id and self.fiscalyear_id.id or False
        result['journal_ids'] = self.journal_ids and self.journal_ids or False
        result['chart_account_id'] = self.chart_account_id and self.chart_account_id.id or False
        result['state'] = self.target_move and self.target_move or ''
        if self.filter == 'filter_date':
            result['date_from'] = self.date_from
            result['date_to'] = self.date_to
        elif self.filter == 'filter_period':
            if not self.period_from or not self.period_to:
                raise osv.except_osv(_('Error!'),_('Select a starting and an ending period.'))
            result['period_from'] = self.period_from
            result['period_to'] = self.period_to
        result['lang'] = self._context.get('lang','en_US')
        return result


    def Contextget_lines(self):        
        used_context = self._build_contexts()
        lines = []
        ids2 = self.env['account.financial.report'].with_context(used_context)._get_children_by_order([self.account_report_id.id])
        for report in self.env['account.financial.report'].with_context(user_context).browse(ids2):
            vals = {
                'name': report.name,
                'balance': report.balance * report.sign or 0.0,
                'type': 'report',
                'level': bool(report.style_overwrite) and report.style_overwrite or report.level,
                'account_type': report.type =='sum' and 'view' or False, #used to underline the financial report balances
            }
            if self.debit_credit:
                vals['debit'] = report.debit
                vals['credit'] = report.credit
            if self.enable_filter:
                vals['balance_cmp'] = self.env['account.financial.report'].with_context(self.comparison_context).browse(report.id).balance * report.sign or 0.0
            lines.append(vals)
            account_ids = []
            if report.display_detail == 'no_detail':
                #the rest of the loop is used to display the details of the financial report, so it's not needed here.
                continue
            if report.type == 'accounts' and report.account_ids:
                account_ids = self.env['account.account']._get_children_and_consol([x.id for x in report.account_ids])
            elif report.type == 'account_type' and report.account_type_ids:
                account_ids = self.env['account.account'].search([('user_type','in', [x.id for x in report.account_type_ids])])
            if account_ids:
                for account in self.env['account.account'].browse(account_ids, context=used_context):
                    #if there are accounts to display, we add them to the lines with a level equals to their level in
                    #the COA + 1 (to avoid having them with a too low level that would conflicts with the level of data
                    #financial reports for Assets, liabilities...)
                    if report.display_detail == 'detail_flat' and account.type == 'view':
                        continue
                    flag = False
                    vals = {
                        'name': account.code + ' ' + account.name,
                        'balance':  account.balance != 0 and account.balance * report.sign or account.balance,
                        'type': 'account',
                        'level': report.display_detail == 'detail_with_hierarchy' and min(account.level + 1,6) or 6, #account.level + 1
                        'account_type': account.type,
                    }

                    if self.debit_credit:
                        vals['debit'] = account.debit
                        vals['credit'] = account.credit
                    if not currency_obj.is_zero(cr, uid, account.company_id.currency_id, vals['balance']):
                        flag = True
                    if self.enable_filter:
                        vals['balance_cmp'] = self.env['account.account'].with_context(self.comparison_context).browse(account.id).balance * report.sign or 0.0
                        if not self.env['res.currency'].is_zero(account.company_id.currency_id.id, vals['balance_cmp']):
                            flag = True
                    if flag:
                        lines.append(vals)
        return lines


 
    #~ def get_lines(self, cr,uid, data):
        #~ lines = []
        #~ account_obj = self.pool.get('account.account')
        #~ currency_obj = self.pool.get('res.currency')
        #~ ids2 = self.pool.get('account.financial.report')._get_children_by_order(cr, uid, [data['form']['account_report_id'][0]], context=data['form']['used_context'])
        #~ for report in self.pool.get('account.financial.report').browse(cr, uid, ids2, context=data['form']['used_context']):
            #~ vals = {
                #~ 'name': report.name,
                #~ 'balance': report.balance * report.sign or 0.0,
                #~ 'type': 'report',
                #~ 'level': bool(report.style_overwrite) and report.style_overwrite or report.level,
                #~ 'account_type': report.type =='sum' and 'view' or False, #used to underline the financial report balances
            #~ }
            #~ if data['form']['debit_credit']:
                #~ vals['debit'] = report.debit
                #~ vals['credit'] = report.credit
            #~ if data['form']['enable_filter']:
                #~ vals['balance_cmp'] = self.pool.get('account.financial.report').browse(self.cr, self.uid, report.id, context=data['form']['comparison_context']).balance * report.sign or 0.0
            #~ lines.append(vals)
            #~ account_ids = []
            #~ if report.display_detail == 'no_detail':
                #~ #the rest of the loop is used to display the details of the financial report, so it's not needed here.
                #~ continue
            #~ if report.type == 'accounts' and report.account_ids:
                #~ account_ids = account_obj._get_children_and_consol(cr, uid, [x.id for x in report.account_ids])
            #~ elif report.type == 'account_type' and report.account_type_ids:
                #~ account_ids = account_obj.search(cr, uid, [('user_type','in', [x.id for x in report.account_type_ids])])
            #~ if account_ids:
                #~ for account in account_obj.browse(cr, uid, account_ids, context=data['form']['used_context']):
                    #~ #if there are accounts to display, we add them to the lines with a level equals to their level in
                    #~ #the COA + 1 (to avoid having them with a too low level that would conflicts with the level of data
                    #~ #financial reports for Assets, liabilities...)
                    #~ if report.display_detail == 'detail_flat' and account.type == 'view':
                        #~ continue
                    #~ flag = False
                    #~ vals = {
                        #~ 'name': account.code + ' ' + account.name,
                        #~ 'balance':  account.balance != 0 and account.balance * report.sign or account.balance,
                        #~ 'type': 'account',
                        #~ 'level': report.display_detail == 'detail_with_hierarchy' and min(account.level + 1,6) or 6, #account.level + 1
                        #~ 'account_type': account.type,
                    #~ }

                    #~ if data['form']['debit_credit']:
                        #~ vals['debit'] = account.debit
                        #~ vals['credit'] = account.credit
                    #~ if not currency_obj.is_zero(cr, uid, account.company_id.currency_id, vals['balance']):
                        #~ flag = True
                    #~ if data['form']['enable_filter']:
                        #~ vals['balance_cmp'] = account_obj.browse(cr, uid, account.id, context=data['form']['comparison_context']).balance * report.sign or 0.0
                        #~ if not currency_obj.is_zero(cr, uid, account.company_id.currency_id, vals['balance_cmp']):
                            #~ flag = True
                    #~ if flag:
                        #~ lines.append(vals)
        #~ return lines


    #~ def get_lines(self, data):
        #~ lines = []
        #~ ids2 = self.env['account.financial.report']._get_children_by_order([data['form']['account_report_id'][0]], context=data['form']['used_context'])
        #~ for report in self.env['account.financial.report'].browse(ids2, context=data['form']['used_context']):
            #~ vals = {
                #~ 'name': report.name,
                #~ 'balance': report.balance * report.sign or 0.0,
                #~ 'type': 'report',
                #~ 'level': bool(report.style_overwrite) and report.style_overwrite or report.level,
                #~ 'account_type': report.type =='sum' and 'view' or False, #used to underline the financial report balances
            #~ }
            #~ if data['form']['debit_credit']:
                #~ vals['debit'] = report.debit
                #~ vals['credit'] = report.credit
            #~ if data['form']['enable_filter']:
                #~ vals['balance_cmp'] = self.pool.get('account.financial.report').browse(self.cr, self.uid, report.id, context=data['form']['comparison_context']).balance * report.sign or 0.0
            #~ lines.append(vals)
            #~ account_ids = []
            #~ if report.display_detail == 'no_detail':
                #~ #the rest of the loop is used to display the details of the financial report, so it's not needed here.
                #~ continue
            #~ if report.type == 'accounts' and report.account_ids:
                #~ account_ids = self.env['account.account']._get_children_and_consol([x.id for x in report.account_ids])
            #~ elif report.type == 'account_type' and report.account_type_ids:
                #~ account_ids = self.env['account.account'].search([('user_type','in', [x.id for x in report.account_type_ids])],context=data['form']['used_context'])
            #~ if account_ids:
                #~ for account in account_ids:
                    #~ #if there are accounts to display, we add them to the lines with a level equals to their level in
                    #~ #the COA + 1 (to avoid having them with a too low level that would conflicts with the level of data
                    #~ #financial reports for Assets, liabilities...)
                    #~ if report.display_detail == 'detail_flat' and account.type == 'view':
                        #~ continue
                    #~ flag = False
                    #~ vals = {
                        #~ 'name': account.code + ' ' + account.name,
                        #~ 'balance':  account.balance != 0 and account.balance * report.sign or account.balance,
                        #~ 'type': 'account',
                        #~ 'level': report.display_detail == 'detail_with_hierarchy' and min(account.level + 1,6) or 6, #account.level + 1
                        #~ 'account_type': account.type,
                    #~ }

                    #~ if data['form']['debit_credit']:
                        #~ vals['debit'] = account.debit
                        #~ vals['credit'] = account.credit
                    #~ if not self.env['res.currency'].is_zero(account.company_id.currency_id, vals['balance']):
                        #~ flag = True
                    #~ if data['form']['enable_filter']:
                        #~ vals['balance_cmp'] = self.env['account.account'].browse(account.id, context=data['form']['comparison_context']).balance * report.sign or 0.0
                        #~ if not self.env['res.currency'].is_zero(account.company_id.currency_id, vals['balance_cmp']):
                            #~ flag = True
                    #~ if flag:
                        #~ lines.append(vals)
        #~ return lines




# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
